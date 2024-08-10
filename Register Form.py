import openpyxl as px
import tkinter as tk

# Load the existing Excel file

workbook = px.load_workbook('PATH') #EXAMPLE: C:\\Users\\Admin\\regis.xlsx
sheet = workbook.active

def initialize_excel():
    # Set column widths and headers
    columns = {
        'A': (30, "Full Name"),
        'B': (15, "Program"),
        'C': (15, "Term"),
        'D': (25, "Application ID"),
        'E': (25, "Phone Number"),
        'F': (35, "Email Address"),
        'G': (45, "Home Address")
    }


    for col, (width, header) in columns.items():
        sheet.column_dimensions[col].width = width
        sheet[f'{col}1'] = header



def focus_next(event, next_widget):
    next_widget.focus_set()

def clear_entries():
    for entry in entries:
        entry.delete(0, tk.END)

def submit_data():
    if all(entry.get() == "" for entry in entries):

        print("Please fill in all fields")

    else:
        new_row = sheet.max_row + 1

        for col, entry in enumerate(entries, start=1):

            sheet.cell(row=new_row, column=col).value = entry.get()
        workbook.save('PATH') #same path
        entries[0].focus_set()
        clear_entries()

# Create the main window

window = tk.Tk()

window.configure(background='light yellow')

window.title("Student Registration")

window.geometry("550x350")

initialize_excel()

# Create labels and entry fields

labels_text = ["Full Name", "Program", "Term", "Application ID", "Phone Number", "Email Address", "Home Address"]
entries = []

for i, text in enumerate(labels_text):
    tk.Label(window, text=text, bg="light yellow").grid(row=i, column=0, padx=10, pady=5)
    entry = tk.Entry(window)
    entry.grid(row=i, column=1, ipadx="100", padx=10, pady=5)
    entries.append(entry)
    if i > 0:
        entries[i-1].bind("<Return>", lambda event, next_widget=entry: focus_next(event, next_widget))



# Create the Submit button


submit_btn = tk.Button(window, text="Submit", fg="White", bg="Green", command=submit_data)
submit_btn.grid(row=len(labels_text), column=1, pady=20)


# Start the GUI event loop
window.mainloop()
